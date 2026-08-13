"""History endpoints: list, get, delete past analyses; PDF/CSV/XLSX export."""
import csv
import io

from fastapi import APIRouter, Depends, Query
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database.session import get_db
from app.schemas.analysis import AnalysisListItem, AnalysisResponse, PaginatedAnalyses
from app.services.analysis_service import AnalysisService
from app.services.pdf_service import generate_analysis_pdf

router = APIRouter(tags=["History"])


def _autosize_column(ws, col_index: int, values: list, min_width: int = 12, max_width: int = 60, padding: int = 2) -> None:
    """Auto-calculate a sensible Excel column width from header + content length.

    XLSX (unlike plain CSV) stores column-width metadata, so this is what
    actually prevents Excel from rendering narrow numeric/date columns as
    ###### -- the root cause of the reported spreadsheet bug.
    """
    longest = max([len(str(v)) for v in values] + [0])
    ws.column_dimensions[get_column_letter(col_index)].width = min(max(longest + padding, min_width), max_width)


@router.get("/history", response_model=PaginatedAnalyses)
def list_history(
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    classification: str | None = None,
    risk_level: str | None = None,
    search: str | None = Query(default=None, description="Full-text search over summary and input text."),
    sort_by: str = Query(default="created_at", pattern="^(created_at|confidence|risk_score|classification|risk_level)$"),
    sort_order: str = Query(default="desc", pattern="^(asc|desc)$"),
    db: Session = Depends(get_db),
) -> PaginatedAnalyses:
    service = AnalysisService(db)
    items, total = service.list_paginated(page, page_size, classification, risk_level, search, sort_by, sort_order)
    return PaginatedAnalyses(
        total=total,
        page=page,
        page_size=page_size,
        items=[AnalysisListItem.model_validate(i) for i in items],
    )


@router.get("/history/export/csv")
def export_history_csv(
    classification: str | None = None,
    risk_level: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
) -> Response:
    items = AnalysisService(db).list_all_matching(classification, risk_level, search)

    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(
        ["Incident ID", "Timestamp", "Classification", "Severity", "Confidence", "Risk Score", "Summary"]
    )
    for item in items:
        writer.writerow(
            [
                item.id,
                item.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                item.classification,
                item.risk_level,
                f"{item.confidence}%",
                item.risk_score if item.risk_score is not None else "",
                item.summary,
            ]
        )

    headers = {"Content-Disposition": 'attachment; filename="security-triage-history.csv"'}
    return Response(content=buffer.getvalue(), media_type="text/csv", headers=headers)


@router.get("/history/export/xlsx")
def export_history_xlsx(
    classification: str | None = None,
    risk_level: str | None = None,
    search: str | None = None,
    db: Session = Depends(get_db),
) -> Response:
    """Real XLSX export with computed column widths, frozen header, autofilter,
    and proper date/numeric cell types -- fixes the Excel '######' rendering
    bug, which is a genuine XLSX column-width problem and cannot be solved via
    plain CSV (CSV carries no width/formatting metadata).
    """
    items = AnalysisService(db).list_all_matching(classification, risk_level, search)

    wb = Workbook()
    ws = wb.active
    ws.title = "Security Triage History"

    headers_row = ["Incident ID", "Timestamp", "Classification", "Severity", "Confidence", "Risk Score", "Summary"]
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers_row, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    col_text_values: dict[int, list[str]] = {i: [] for i in range(1, len(headers_row) + 1)}

    for row_idx, item in enumerate(items, start=2):
        ws.cell(row=row_idx, column=1, value=item.id)

        # Store a real datetime value (not a string) with a readable Excel
        # date/time number format, and give the column enough width for it --
        # this combination is exactly what plain CSV cannot express.
        created_at = item.created_at.replace(tzinfo=None) if item.created_at.tzinfo else item.created_at
        ts_cell = ws.cell(row=row_idx, column=2, value=created_at)
        ts_cell.number_format = "yyyy-mm-dd hh:mm"

        ws.cell(row=row_idx, column=3, value=item.classification)
        ws.cell(row=row_idx, column=4, value=item.risk_level)

        conf_cell = ws.cell(row=row_idx, column=5, value=(item.confidence or 0) / 100)
        conf_cell.number_format = "0%"
        conf_cell.alignment = Alignment(horizontal="center")

        score_cell = ws.cell(row=row_idx, column=6, value=item.risk_score if item.risk_score is not None else None)
        score_cell.number_format = "0"
        score_cell.alignment = Alignment(horizontal="center")

        summary_cell = ws.cell(row=row_idx, column=7, value=item.summary)
        summary_cell.alignment = Alignment(wrap_text=True, vertical="top")

        col_text_values[1].append(item.id)
        col_text_values[2].append(created_at.strftime("%Y-%m-%d %H:%M"))
        col_text_values[3].append(item.classification)
        col_text_values[4].append(item.risk_level)
        col_text_values[5].append(f"{item.confidence}%")
        col_text_values[6].append(str(item.risk_score) if item.risk_score is not None else "")
        col_text_values[7].append(item.summary or "")

    last_row = ws.max_row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers_row))}{max(last_row, 1)}"

    max_widths = {1: 38, 2: 20, 3: 24, 4: 14, 5: 12, 6: 12, 7: 60}
    min_widths = {1: 20, 2: 18, 3: 12, 4: 12, 5: 12, 6: 12, 7: 30}
    for col_idx, header in enumerate(headers_row, start=1):
        values = [header] + col_text_values[col_idx]
        _autosize_column(ws, col_idx, values, min_width=min_widths[col_idx], max_width=max_widths[col_idx])

    buffer = io.BytesIO()
    wb.save(buffer)
    xlsx_headers = {"Content-Disposition": 'attachment; filename="security-triage-history.xlsx"'}
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=xlsx_headers,
    )


@router.get("/history/{analysis_id}", response_model=AnalysisResponse)
def get_history_item(analysis_id: str, db: Session = Depends(get_db)) -> AnalysisResponse:
    analysis = AnalysisService(db).get_or_404(analysis_id)
    return AnalysisResponse.model_validate(analysis)


@router.delete("/history/{analysis_id}", status_code=204)
def delete_history_item(analysis_id: str, db: Session = Depends(get_db)) -> None:
    AnalysisService(db).delete(analysis_id)


@router.get("/history/{analysis_id}/export")
def export_history_item(analysis_id: str, db: Session = Depends(get_db)) -> Response:
    analysis = AnalysisService(db).get_or_404(analysis_id)
    pdf_bytes = generate_analysis_pdf(analysis)
    headers = {"Content-Disposition": f'attachment; filename="analysis-{analysis_id}.pdf"'}
    return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)
